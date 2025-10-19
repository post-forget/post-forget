import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os

import argparse

parser = argparse.ArgumentParser(description="Generate pivot tables from model comparison data.")
parser.add_argument(
    '--input-file',
    type=str,
    default="merged_file.csv",
    help="Path to the input CSV file (default: merged_file.csv)"
)
parser.add_argument(
    '--output-prefix',
    type=str,
    default="ReasoningForgettingLM_Data",
    help="Prefix for the output XLSX and CSV files (default: ReasoningForgettingLM_Data)"
)
parser.add_argument(
    '--pivot-dir',
    type=str,
    default="pivot_tables_csv",
    help="Directory to save the individual pivot table CSV files (default: pivot_tables_csv)"
)

args = parser.parse_args()

output_excel_path = f"{args.output_prefix}.xlsx"
output_csv_path = f"{args.output_prefix}.csv"
pivot_dir_path = args.pivot_dir

print(f"Loading data from {args.input_file}...")
df = pd.read_csv(args.input_file)

num_choices_map = {
    "community|bbh:boolean_expressions|0": 2,
    "community|bbh:causal_judgment|0": 2,
    "community|bbh:date_understanding|0": 6,
    "community|bbh:disambiguation_qa|0": 3,
    "community|bbh:dyck_languages|0": 0,
    "community|bbh:formal_fallacies|0": 2,
    "community|bbh:geometric_shapes|0": 11,
    "community|bbh:hyperbaton|0": 2,
    "community|bbh:logical_deduction_five_objects|0": 5,
    "community|bbh:logical_deduction_seven_objects|0": 7,
    "community|bbh:logical_deduction_three_objects|0": 3,
    "community|bbh:movie_recommendation|0": 6,
    "community|bbh:multistep_arithmetic_two|0": 0,
    "community|bbh:navigate|0": 2,
    "community|bbh:object_counting|0": 19,
    "community|bbh:penguins_in_a_table|0": 5,
    "community|bbh:reasoning_about_colored_objects|0": 18,
    "community|bbh:ruin_names|0": 6,
    "community|bbh:salient_translation_error_detection|0": 6,
    "community|bbh:snarks|0": 2,
    "community|bbh:sports_understanding|0": 2,
    "community|bbh:temporal_sequences|0": 4,
    "community|bbh:tracking_shuffled_objects_five_objects|0": 5,
    "community|bbh:tracking_shuffled_objects_seven_objects|0": 7,
    "community|bbh:tracking_shuffled_objects_three_objects|0": 3,
    "community|bbh:web_of_lies|0": 2,
    "community|bbh:word_sorting|0": 0,
}

# ignore these tasks
explicit_ignore = {
    "community|bbh:dyck_languages|0",
    "community|truthfulqa:mc2|0",
    "community|bbh:word_sorting|0",
    "community|bbh:multistep_arithmetic_two|0",
}

# and ignore global_mmlu results
for t in df["task"].unique():
    if "global_mmlu" in str(t):
        explicit_ignore.add(t)

df["Ignore"] = df["task"].isin(explicit_ignore).astype(int)

df["num_choices"] = df["task"].map(num_choices_map).fillna(4).astype(int)

# ---------------------------
# Category mapping
# ---------------------------
category_map = [
    ("commonsense", "community|commonsense_qa:main|0"),
    ("commonsense", "community|piqa:main|0"),
    ("culture", "community|bbh:sports_understanding|0"),
    ("culture", "community|bbh:movie_recommendation|0"),
    ("deduction", "community|bbh:navigate|0"),
    ("deduction", "community|bbh:causal_judgment|0"),
    ("deduction", "community|bbh:penguins_in_a_table|0"),
    ("deduction", "community|bbh:web_of_lies|0"),
    ("deduction", "community|bbh:tracking_shuffled_objects_three_objects|0"),
    ("deduction", "community|bbh:tracking_shuffled_objects_seven_objects|0"),
    ("deduction", "community|bbh:tracking_shuffled_objects_five_objects|0"),
    ("deduction", "community|bbh:temporal_sequences|0"),
    ("deduction", "community|bbh:reasoning_about_colored_objects|0"),
    ("deduction", "community|bbh:logical_deduction_three_objects|0"),
    ("deduction", "community|bbh:logical_deduction_seven_objects|0"),
    ("deduction", "community|bbh:logical_deduction_five_objects|0"),
    ("deduction", "community|bbh:formal_fallacies|0"),
    ("deduction", "community|bbh:date_understanding|0"),
    ("deduction", "community|arc:easy|0"),
    ("deduction", "community|arc:challenge|0"),
    ("deduction", "community|musr:murder_mysteries|0"),
    ("deduction", "community|musr:object_placements|0"),
    ("deduction", "community|musr:team_allocation|0"),
    ("deduction", "community|mmlu:logical_fallacies|0"),
    ("ignore", "community|bbh:dyck_languages|0"),
    ("ignore", "community|truthfulqa:mc2|0"),
    ("knowledge_and_qa", "community|bbh:object_counting|0"),
    ("knowledge_and_qa", "community|mmlu:miscellaneous|0"),
    ("knowledge_and_qa", "community|mmlu:global_facts|0"),
    ("knowledge_and_qa", "community|mctest:main|0"),
    ("language_and_communication", "community|bbh:snarks|0"),
    ("language_and_communication", "community|bbh:disambiguation_qa|0"),
    ("ignore", "community|bbh:word_sorting|0"),
    ("language_and_communication", "community|bbh:ruin_names|0"),
    ("language_and_communication", "community|bbh:hyperbaton|0"),
    ("language_and_communication", "community|social_iqa:main|0"),
    ("language_and_communication", "community|hellaswag:main|0"),
    ("language_and_communication", "community|bbh:salient_translation_error_detection|0"),
    ("liberal_arts", "community|mmlu:world_religions|0"),
    ("liberal_arts", "community|mmlu:us_foreign_policy|0"),
    ("liberal_arts", "community|mmlu:sociology|0"),
    ("liberal_arts", "community|mmlu:security_studies|0"),
    ("liberal_arts", "community|mmlu:public_relations|0"),
    ("liberal_arts", "community|mmlu:professional_psychology|0"),
    ("liberal_arts", "community|mmlu:professional_law|0"),
    ("liberal_arts", "community|mmlu:prehistory|0"),
    ("liberal_arts", "community|mmlu:philosophy|0"),
    ("liberal_arts", "community|mmlu:management|0"),
    ("liberal_arts", "community|mmlu:international_law|0"),
    ("liberal_arts", "community|mmlu:high_school_world_history|0"),
    ("liberal_arts", "community|mmlu:high_school_us_history|0"),
    ("liberal_arts", "community|mmlu:high_school_psychology|0"),
    ("liberal_arts", "community|mmlu:high_school_microeconomics|0"),
    ("liberal_arts", "community|mmlu:high_school_macroeconomics|0"),
    ("liberal_arts", "community|mmlu:high_school_government_and_politics|0"),
    ("liberal_arts", "community|mmlu:high_school_geography|0"),
    ("liberal_arts", "community|mmlu:high_school_european_history|0"),
    ("ignore", "community|bbh:multistep_arithmetic_two|0"),
    ("math", "community|bbh:geometric_shapes|0"),
    ("math", "community|bbh:boolean_expressions|0"),
    ("math", "community|mmlu:high_school_statistics|0"),
    ("math", "community|mmlu:high_school_mathematics|0"),
    ("math", "community|mmlu:formal_logic|0"),
    ("math", "community|mmlu:elementary_mathematics|0"),
    ("math", "community|mmlu:econometrics|0"),
    ("math", "community|mmlu:college_mathematics|0"),
    ("math", "community|mmlu:abstract_algebra|0"),
    ("safety_and_truthfulness", "community|mmlu:moral_scenarios|0"),
    ("safety_and_truthfulness", "community|mmlu:moral_disputes|0"),
    ("safety_and_truthfulness", "community|mmlu:jurisprudence|0"),
    ("safety_and_truthfulness", "community|mmlu:business_ethics|0"),
    ("safety_and_truthfulness", "community|truthfulqa:mc1|0"),
    ("safety_and_truthfulness", "community|salad_bench:mrq|0"),
    ("instruction_following", "community|ifeval:main|0"),
    ("science,technology,engineering", "community|mmlu:marketing|0"),
    ("science,technology,engineering", "community|mmlu:virology|0"),
    ("science,technology,engineering", "community|mmlu:professional_medicine|0"),
    ("science,technology,engineering", "community|mmlu:professional_accounting|0"),
    ("science,technology,engineering", "community|mmlu:nutrition|0"),
    ("science,technology,engineering", "community|mmlu:medical_genetics|0"),
    ("science,technology,engineering", "community|mmlu:machine_learning|0"),
    ("science,technology,engineering", "community|mmlu:human_sexuality|0"),
    ("science,technology,engineering", "community|mmlu:human_aging|0"),
    ("science,technology,engineering", "community|mmlu:high_school_physics|0"),
    ("science,technology,engineering", "community|mmlu:high_school_computer_science|0"),
    ("science,technology,engineering", "community|mmlu:high_school_chemistry|0"),
    ("science,technology,engineering", "community|mmlu:high_school_biology|0"),
    ("science,technology,engineering", "community|mmlu:electrical_engineering|0"),
    ("science,technology,engineering", "community|mmlu:conceptual_physics|0"),
    ("science,technology,engineering", "community|mmlu:computer_security|0"),
    ("science,technology,engineering", "community|mmlu:college_physics|0"),
    ("science,technology,engineering", "community|mmlu:college_medicine|0"),
    ("science,technology,engineering", "community|mmlu:college_computer_science|0"),
    ("science,technology,engineering", "community|mmlu:college_chemistry|0"),
    ("science,technology,engineering", "community|mmlu:college_biology|0"),
    ("science,technology,engineering", "community|mmlu:clinical_knowledge|0"),
    ("science,technology,engineering", "community|mmlu:astronomy|0"),
    ("science,technology,engineering", "community|mmlu:anatomy|0"),
    ("science,technology,engineering", "community|gpqa:diamond|0"),
]

# Normalize and create category mapping
category_dict = {task: cat for cat, task in category_map}
df["task"] = df["task"].astype(str).str.strip()
df["Category"] = df["task"].map(category_dict).fillna("Missing")

df["Number of Correct Answers"] = df["task"].apply(
    lambda t: None if t in ["community|truthfulqa:mc2|0", "community|truthfulqa_reasoning_letter:mc2|0"] else 1
)

df["Abs. Forget."] = df["1->0"] / df["n"]
df["Rel. Forget."] = df["1->0"] / (df["n"] * df["acc_before"])
df["Abs. Improvement"] = df["0->1"] / df["n"]
df["Rel. Improvement"] = df["0->1"] / (df["n"] * df["acc_before"])

denominator = df["num_choices"] - 1
adjustment_term = ((1 - df["acc_before"]) * (1 - df["acc_after"])) / denominator
df["Adj. Abs. Forgetting (Simplified)"] = (df["Abs. Forget."] - adjustment_term).clip(lower=0)
df["Adj. Abs. Improvement (Simplified)"] = (df["Abs. Improvement"] - adjustment_term).clip(lower=0)

# Copy to main adjusted columns
df["Adj. Abs. Forgetting"] = df["Adj. Abs. Forgetting (Simplified)"].copy()
df["Adj. Abs. Improvement"] = df["Adj. Abs. Improvement (Simplified)"].copy()

# Adj. 2 Seed metrics
denom_2seed = (
    df["mean_count_2_seeds_0_0"]
    + df["mean_count_2_seeds_0_1"]
    + df["mean_count_2_seeds_1_0"]
    + df["mean_count_2_seeds_1_1"]
)
df["Adj. 2 Seed Abs. Forgetting"] = df["mean_count_2_seeds_1_0"] / denom_2seed
df["Adj. 2 Seed Abs. Improvement"] = df["mean_count_2_seeds_0_1"] / denom_2seed


df["Max Possible Forgetting"] = np.where(
    df["num_choices"] - 1 > 0,
    np.maximum((df["num_choices"] * df["acc_before"] - 1) / (df["num_choices"] - 1), 0),
    np.nan,
)

df["Max Possible Improvement"] = np.where(
    df["num_choices"] - 1 > 0,
    np.maximum((df["num_choices"] * df["acc_after"] - 1) / (df["num_choices"] - 1), 0),
    np.nan,
)


model_groups = {
    "Instruction Tuning (All)": [
        ("Qwen/Qwen2.5-3B", "Qwen_Qwen2.5-B-Instruct"),
        ("Qwen/Qwen2.5-3B", "Qwen/Qwen2.5-3B-Instruct"),
        ("Qwen/Qwen2.5-7B", "Qwen_Qwen2.5-7B-Instruct"),
        ("Qwen/Qwen2.5-7B", "Qwen/Qwen2.5-7B-Instruct"),
        ("Qwen/Qwen2.5-14B", "Qwen_Qwen2.5-14B-Instruct"),
        ("Qwen/Qwen2.5-14B", "Qwen/Qwen2.5-14B-Instruct"),
        ("Qwen/Qwen2.5-32B", "Qwen_Qwen2.5-32B-Instruct"),
        ("Qwen/Qwen2.5-32B", "Qwen/Qwen2.5-32B-Instruct"),

        ("Qwen_Qwen2.5-3B", "Qwen_Qwen2.5-3B-Instruct"),
        ("Qwen_Qwen2.5-3B", "Qwen/Qwen2.5-3B-Instruct"),
        ("Qwen_Qwen2.5-7B", "Qwen_Qwen2.5-7B-Instruct"),
        ("Qwen_Qwen2.5-7B", "Qwen/Qwen2.5-7B-Instruct"),
        ("Qwen_Qwen2.5-14B", "Qwen_Qwen2.5-14B-Instruct"),
        ("Qwen_Qwen2.5-14B", "Qwen/Qwen2.5-14B-Instruct"),
        ("Qwen_Qwen2.5-32B", "Qwen_Qwen2.5-32B-Instruct"),
        ("Qwen_Qwen2.5-32B", "Qwen/Qwen2.5-32B-Instruct"),

        ("meta-llama/Llama-3.1-8B", "meta-llama_Llama-3.1-8B-Instruct"),

        ("Qwen/Qwen2.5-Coder-3B", "Qwen/Qwen2.5-Coder-3B-Instruct"),
        ("Qwen/Qwen2.5-Coder-3B", "Qwen_Qwen2.5-Coder-3B-Instruct"),
        ("Qwen/Qwen2.5-Coder-7B", "Qwen_Qwen2.5-Coder-7B-Instruct"),
        ("Qwen/Qwen2.5-Coder-7B", "Qwen/Qwen2.5-Coder-7B-Instruct"),
        ("Qwen/Qwen2.5-Coder-14B", "Qwen_Qwen2.5-Coder-14B-Instruct"),
        ("Qwen/Qwen2.5-Coder-14B", "Qwen/Qwen2.5-Coder-14B-Instruct"),
        ("Qwen/Qwen2.5-Coder-32B", "Qwen_Qwen2.5-Coder-32B-Instruct"),
        ("Qwen/Qwen2.5-Coder-32B", "Qwen/Qwen2.5-Coder-32B-Instruct"),
        ("Qwen/Qwen2.5-Math-7B", "Qwen/Qwen2.5-7B-Math-Instruct"),

        ("Qwen_Qwen2.5-Coder-3B", "Qwen/Qwen2.5-Coder-3B-Instruct"),
        ("Qwen_Qwen2.5-Coder-3B", "Qwen_Qwen2.5-Coder-3B-Instruct"),
        ("Qwen_Qwen2.5-Coder-7B", "Qwen_Qwen2.5-Coder-7B-Instruct"),
        ("Qwen_Qwen2.5-Coder-7B", "Qwen/Qwen2.5-Coder-7B-Instruct"),
        ("Qwen_Qwen2.5-Coder-14B", "Qwen_Qwen2.5-Coder-14B-Instruct"),
        ("Qwen_Qwen2.5-Coder-14B", "Qwen/Qwen2.5-Coder-14B-Instruct"),
        ("Qwen_Qwen2.5-Coder-32B", "Qwen_Qwen2.5-Coder-32B-Instruct"),
        ("Qwen_Qwen2.5-Coder-32B", "Qwen/Qwen2.5-Coder-32B-Instruct"),
        ("Qwen_Qwen2.5-Math-7B", "Qwen/Qwen2.5-7B-Math-Instruct"),
    ],
    "Task Training (All)": [
        ("Qwen/Qwen2.5-3B", "Qwen/Qwen2.5-Coder-3B"),
        ("Qwen/Qwen2.5-7B", "Qwen/Qwen2.5-Coder-7B"),
        ("Qwen/Qwen2.5-14B", "Qwen/Qwen2.5-Coder-14B"),
        ("Qwen/Qwen2.5-32B", "Qwen/Qwen2.5-Coder-32B"),

        ("Qwen_Qwen2.5-3B", "Qwen/Qwen2.5-Coder-3B"),
        ("Qwen_Qwen2.5-7B", "Qwen/Qwen2.5-Coder-7B"),
        ("Qwen_Qwen2.5-14B", "Qwen/Qwen2.5-Coder-14B"),
        ("Qwen_Qwen2.5-32B", "Qwen/Qwen2.5-Coder-32B"),

        ("Qwen/Qwen2.5-3B", "Qwen_Qwen2.5-Coder-3B"),
        ("Qwen/Qwen2.5-7B", "Qwen_Qwen2.5-Coder-7B"),
        ("Qwen/Qwen2.5-14B", "Qwen_Qwen2.5-Coder-14B"),
        ("Qwen/Qwen2.5-32B", "Qwen_Qwen2.5-Coder-32B"),

        ("Qwen_Qwen2.5-3B", "Qwen_Qwen2.5-Coder-3B"),
        ("Qwen_Qwen2.5-7B", "Qwen_Qwen2.5-Coder-7B"),
        ("Qwen_Qwen2.5-14B", "Qwen_Qwen2.5-Coder-14B"),
        ("Qwen_Qwen2.5-32B", "Qwen_Qwen2.5-Coder-32B"),

        ("Qwen_Qwen2.5-7B", "Qwen_Qwen2.5-Math-7B"),
        ("Qwen_Qwen2.5-7B", "Qwen/Qwen2.5-Math-7B"),
        ("Qwen/Qwen2.5-7B", "Qwen_Qwen2.5-Math-7B"),
        ("Qwen/Qwen2.5-7B", "Qwen/Qwen2.5-Math-7B"),
    ],
    "Trained from Base (All)": [
        ("Qwen/Qwen2.5-Math-7B", "Qwen/Qwen2.5-Math-7B-Instruct"),
        ("Qwen/Qwen2.5-Math-7B", "Qwen_Qwen2.5-Math-7B-Instruct"),
        ("Qwen_Qwen2.5-Math-7B", "Qwen/Qwen2.5-Math-7B-Instruct"),
        ("Qwen_Qwen2.5-Math-7B", "Qwen_Qwen2.5-Math-7B-Instruct"),

        ("Qwen/Qwen2.5-32B", "Qwen/QwQ-32B"),
        ("Qwen_Qwen2.5-32B", "Qwen/QwQ-32B"),

        ("Qwen/Qwen2.5-Math-7B", "deepseek-ai_DeepSeek-R1-Distill-Qwen-7B"),
        ("Qwen_Qwen2.5-Math-7B", "deepseek-ai_DeepSeek-R1-Distill-Qwen-7B"),

        ("meta-llama/Llama-3.1-8B", "deepseek-ai/DeepSeek-R1-Distill-Llama-8B"),
        ("meta-llama/Llama-3.1-8B", "deepseek-ai_DeepSeek-R1-Distill-Llama-8B"),
        ("Qwen/Qwen2.5-14B", "deepseek-ai/DeepSeek-R1-Distill-Qwen-142B"),
        ("Qwen/Qwen2.5-32B", "deepseek-ai/DeepSeek-R1-Distill-Qwen-32B"),

        ("Qwen_Qwen2.5-14B", "deepseek-ai/DeepSeek-R1-Distill-Qwen-142B"),
        ("Qwen_Qwen2.5-32B", "deepseek-ai/DeepSeek-R1-Distill-Qwen-32B"),
    ],
    "Trained from Instruct - SFT (All)": [
        ("Qwen_Qwen2.5-7B-Instruct", "nvidia_OpenReasoning-Nemotron-7B"),
        ("Qwen_Qwen2.5-14B-Instruct", "nvidia_OpenReasoning-Nemotron-14B"),
        ("Qwen_Qwen2.5-32B-Instruct", "nvidia_OpenReasoning-Nemotron-32B"),
        ("Qwen/Qwen2.5-7B-Instruct", "open-thoughts_OpenThinker-7B"),

        ("Qwen/Qwen2.5-7B-Instruct", "open-thoughts/OpenThinker-7B"),
        ("Qwen/Qwen2.5-32B-Instruct", "open-thoughts/OpenThinker2-32B"),
        ("Qwen_Qwen2.5-7B-Instruct", "open-thoughts/OpenThinker-7B"),
        ("Qwen_Qwen2.5-32B-Instruct", "open-thoughts/OpenThinker2-32B"),

        ("Qwen/Qwen2.5-7B-Instruct", "open-thoughts/OpenThinker3-7B"),
        ("Qwen_Qwen2.5-7B-Instruct", "simplescaling_s1.1-7B"),
        ("Qwen/Qwen2.5-7B-Instruct", "simplescaling_s1.1-7B"),
        ("Qwen_Qwen2.5-14B-Instruct", "simplescaling_s1.1-14B"),
        ("Qwen/Qwen2.5-14B-Instruct", "simplescaling_s1.1-14B"),
        ("Qwen_Qwen2.5-32B-Instruct", "simplescaling_s1.1-32B"),
        ("Qwen/Qwen2.5-32B-Instruct", "simplescaling_s1.1-32B"),
    ],
    "Trained from Instruct - Low Data Scenario": [
        ("Qwen_Qwen2.5-7B-Instruct", "simplescaling_s1.1-7B"),
        ("Qwen/Qwen2.5-7B-Instruct", "simplescaling_s1.1-7B"),
        ("Qwen_Qwen2.5-14B-Instruct", "simplescaling_s1.1-14B"),
        ("Qwen/Qwen2.5-14B-Instruct", "simplescaling_s1.1-14B"),
        ("Qwen_Qwen2.5-32B-Instruct", "simplescaling_s1.1-32B"),
        ("Qwen/Qwen2.5-32B-Instruct", "simplescaling_s1.1-32B"),

        ("Qwen/Qwen2.5-32B-Instruct", "GAIR_LIMO"),
        ("Qwen/Qwen2.5-32B-Instruct", "GAIR/LIMO"),

        ("Qwen_Qwen2.5-32B-Instruct", "GAIR_LIMO"),
        ("Qwen_Qwen2.5-32B-Instruct", "GAIR/LIMO"),

        ("Qwen/Qwen2.5-32B-Instruct", "GAIR_LIMO-v2"),
        ("Qwen/Qwen2.5-32B-Instruct", "GAIR/LIMO-v2"),

        ("Qwen_Qwen2.5-32B-Instruct", "GAIR_LIMO-v2"),
        ("Qwen_Qwen2.5-32B-Instruct", "GAIR/LIMO-v2"),
    ],
    "Trained from Instruct - High Data Scenario (All)": [
        ("Qwen/Qwen2.5-7B-Instruct", "nvidia_OpenCodeReasoning-Nemotron-1.1-7B"),
        ("Qwen/Qwen2.5-14B-Instruct", "nvidia_OpenCodeReasoning-Nemotron-1.1-14B"),
        ("Qwen/Qwen2.5-7B-Instruct", "nvidia_OpenCodeReasoning-Nemotron-1.1-32B"),

        ("Qwen_Qwen2.5-7B-Instruct", "nvidia_OpenCodeReasoning-Nemotron-1.1-7B"),
        ("Qwen_Qwen2.5-14B-Instruct", "nvidia_OpenCodeReasoning-Nemotron-1.1-14B"),
        ("Qwen_Qwen2.5-7B-Instruct", "nvidia_OpenCodeReasoning-Nemotron-1.1-32B"),

        ("meta-llama/Llama-3.1-8B", "nvidia/OpenMath2-Llama3.1-8B"),
        ("meta-llama/Llama-3.1-8B", "nvidia_OpenMath2-Llama3.1-8B"),


        ("Qwen/Qwen2.5-7B-Instruct", "open-thoughts/OpenThinker-7B"),
        ("Qwen/Qwen2.5-7B-Instruct", "open-thoughts_OpenThinker-7B"),
        ("Qwen_Qwen2.5-7B-Instruct", "open-thoughts/OpenThinker-7B"),
        ("Qwen_Qwen2.5-7B-Instruct", "open-thoughts_OpenThinker-7B"),

        ("Qwen/Qwen2.5-32B-Instruct", "open-thoughts/OpenThinker2-32B"),
        ("Qwen/Qwen2.5-32B-Instruct", "open-thoughts_OpenThinker2-32B"),
        ("Qwen_Qwen2.5-32B-Instruct", "open-thoughts/OpenThinker2-32B"),
        ("Qwen_Qwen2.5-32B-Instruct", "open-thoughts_OpenThinker2-32B"),

        ("Qwen/Qwen2.5-7B-Instruct", "open-thoughts/OpenThinker3-7B"),
        ("Qwen/Qwen2.5-7B-Instruct", "open-thoughts_OpenThinker3-7B"),
        ("Qwen_Qwen2.5-7B-Instruct", "open-thoughts/OpenThinker3-7B"),
        ("Qwen_Qwen2.5-7B-Instruct", "open-thoughts_OpenThinker3-7B"),


        ("Qwen_QwQ-32B", "PrimeIntellect_INTELLECT-2"),
        ("Qwen_QwQ-32B", "PrimeIntellect/INTELLECT-2"),
        ("Qwen/QwQ-32B", "PrimeIntellect_INTELLECT-2"),
        ("Qwen/QwQ-32B", "PrimeIntellect/INTELLECT-2"),
    ],
}

# Include these metrics in the generated tables
metrics = [
    "Adj. Abs. Forgetting",
    "Adj. Abs. Improvement",
    "acc_before",
    "acc_after",
    "Max Possible Improvement",
    "Max Possible Forgetting"
    # "Adj. 2 Seed Abs. Forgetting",
    # "Adj. 2 Seed Abs. Improvement",
]

def create_pivot_table(df, model_pairs, group_name):
    """
    Strict per-seed-per-category averaging, then compute sample variance (ddof=1)
    and std = sqrt(var) across seeds. Requires 'seed' column.
    Returns pivot_tables, detailed_tables with keys:
      metric, metric + " (var)", metric + " (std)"
    """
    # Filter data for relevant model pairs
    filtered_dfs = []

    for model_before, model_after in model_pairs:
        subset = df[(df["model_before"] == model_before) & (df["model_after"] == model_after)].copy()
        if len(subset) > 0:
            subset["model_pair"] = f"{model_before} → {model_after}"
            filtered_dfs.append(subset)

    if not filtered_dfs:
        print(f"Warning: No data found for {group_name}")
        return None

    combined_df = pd.concat(filtered_dfs, ignore_index=True)

    # Filter out ignored tasks
    combined_df = combined_df[combined_df["Ignore"] == 0]

    if "seed" not in combined_df.columns:
        raise RuntimeError("create_pivot_table requires a 'seed' column in the DataFrame. Aborting.")

    # Debug: Check how many unique seeds we have per model_pair
    print(f"\n{group_name} - Seeds per model pair:")
    seed_counts = combined_df.groupby("model_pair")["seed"].nunique()
    print(seed_counts)

    pivot_tables = {}
    detailed_tables = {}

    for metric in metrics:
        combined_df[metric] = pd.to_numeric(combined_df[metric], errors="coerce")

        # --- per-seed, per-category average (average tasks per seed) ---
        per_seed_cat = (
            combined_df
            .groupby(["model_pair", "seed", "Category"], dropna=False)[metric]
            .mean()
            .reset_index(name="per_seed_cat_mean")
        )

        # compute mean, sample variance (ddof=1) and count across seeds for each (Category, model_pair)
        stats = (
            per_seed_cat
            .groupby(["Category", "model_pair"], dropna=False)["per_seed_cat_mean"]
            .agg(mean="mean", var=lambda x: x.var(ddof=1), count="count")
        )

        # pivot wide
        pivot_mean = stats["mean"].unstack(level="model_pair")
        pivot_var = stats["var"].unstack(level="model_pair")
        pivot_count = stats["count"].unstack(level="model_pair")

        # ensure that positions with count <= 1 have var = NaN (sample variance undefined)
        if pivot_count is not None:
            mask_single = pivot_count <= 1
            if mask_single.any().any():
                pivot_var = pivot_var.mask(mask_single, other=np.nan)

        # compute std from var
        pivot_std = pivot_var.copy().map(lambda v: np.sqrt(v) if pd.notna(v) else np.nan)

        # --- OVERALL: for each seed, mean across categories, then mean/var across seeds ---
        per_seed_overall = (
            per_seed_cat
            .groupby(["model_pair", "seed"], dropna=False)["per_seed_cat_mean"]
            .mean()
            .reset_index(name="per_seed_overall")
        )
        overall_stats = (
            per_seed_overall
            .groupby("model_pair", dropna=False)["per_seed_overall"]
            .agg(mean="mean", var=lambda x: x.var(ddof=1), count="count")
        )

        # Construct OVERALL rows aligned to pivot columns (preserve columns order)
        cols = pivot_mean.columns if (pivot_mean is not None and pivot_mean.shape[1] > 0) else overall_stats.index
        overall_mean_df = pd.DataFrame([overall_stats["mean"].reindex(cols)], index=["OVERALL"])
        overall_var_df = pd.DataFrame([overall_stats["var"].reindex(cols)], index=["OVERALL"])
        overall_count_df = pd.DataFrame([overall_stats["count"].reindex(cols)], index=["OVERALL"])
        overall_var_df = overall_var_df.mask(overall_count_df <= 1, other=np.nan)
        overall_std_df = overall_var_df.copy().map(lambda v: np.sqrt(v) if pd.notna(v) else np.nan)

        # concat, handling empty pivot_mean case
        if pivot_mean is None or pivot_mean.shape[0] == 0:
            pivot_mean = overall_mean_df
        else:
            pivot_mean = pd.concat([pivot_mean, overall_mean_df])

        if pivot_var is None or pivot_var.shape[0] == 0:
            pivot_var = overall_var_df
        else:
            pivot_var = pd.concat([pivot_var, overall_var_df])

        if pivot_std is None or pivot_std.shape[0] == 0:
            pivot_std = overall_std_df
        else:
            pivot_std = pd.concat([pivot_std, overall_std_df])

        pivot_tables[metric] = pivot_mean
        pivot_tables[metric + " (var)"] = pivot_var
        pivot_tables[metric + " (std)"] = pivot_std

        # ---------------- DETAILED (Category, task) mean/var/std across seeds ----------------
        per_seed_task = (
            combined_df
            .groupby(["model_pair", "seed", "Category", "task"], dropna=False)[metric]
            .mean()
            .reset_index(name="per_seed_task_mean")
        )

        detailed_stats = (
            per_seed_task
            .groupby(["Category", "task", "model_pair"], dropna=False)["per_seed_task_mean"]
            .agg(mean="mean", var=lambda x: x.var(ddof=1), count="count")
            .reset_index()
        )

        if not detailed_stats.empty:
            detailed_mean = detailed_stats.pivot(index=["Category", "task"], columns="model_pair", values="mean")
            detailed_var = detailed_stats.pivot(index=["Category", "task"], columns="model_pair", values="var")
            detailed_count = detailed_stats.pivot(index=["Category", "task"], columns="model_pair", values="count")
            # mask single-seed var -> NaN
            if detailed_count is not None:
                detailed_var = detailed_var.mask(detailed_count <= 1, other=np.nan)
            detailed_std = detailed_var.copy().map(lambda v: np.sqrt(v) if pd.notna(v) else np.nan)
        else:
            detailed_mean = pd.DataFrame()
            detailed_var = pd.DataFrame()
            detailed_std = pd.DataFrame()

        detailed_tables[metric] = detailed_mean
        detailed_tables[metric + " (var)"] = detailed_var
        detailed_tables[metric + " (std)"] = detailed_std

    return pivot_tables, detailed_tables


def write_group_sheet(wb, group_name, pivot_tables, detailed_tables):
    """
    Write the pivot summary (mean + std) and detailed (mean + std) tables
    into an Excel sheet in workbook `wb`. Uses the global `metrics` list.
    """
    sheet_name = group_name[:31]  # Excel sheet names limited to 31 chars
    if sheet_name in wb.sheetnames:
        # remove existing to avoid duplicates
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    row_offset = 1

    # Summary tables (means then stds for each metric)
    for metric in metrics:
        # MEAN summary
        pivot_mean = pivot_tables.get(metric)
        if pivot_mean is None:
            continue

        ws.cell(row=row_offset, column=1, value=f"AVERAGE of {metric}")
        row_offset += 1

        # Header row
        ws.cell(row=row_offset, column=1, value="Category")
        for col_idx, col_name in enumerate(pivot_mean.columns, start=2):
            ws.cell(row=row_offset, column=col_idx, value=col_name)
        row_offset += 1

        # Write mean table rows
        for idx_name in pivot_mean.index:
            ws.cell(row=row_offset, column=1, value=idx_name)
            for col_idx, col_name in enumerate(pivot_mean.columns, start=2):
                value = pivot_mean.loc[idx_name, col_name]
                if pd.notna(value):
                    ws.cell(row=row_offset, column=col_idx, value=value)
            row_offset += 1

        row_offset += 1  # small space

        # STD summary (if present)
        std_key = metric + " (std)"
        pivot_std = pivot_tables.get(std_key)
        if pivot_std is not None:
            ws.cell(row=row_offset, column=1, value=f"STD of {metric}")
            row_offset += 1

            ws.cell(row=row_offset, column=1, value="Category")
            for col_idx, col_name in enumerate(pivot_std.columns, start=2):
                ws.cell(row=row_offset, column=col_idx, value=col_name)
            row_offset += 1

            for idx_name in pivot_std.index:
                ws.cell(row=row_offset, column=1, value=idx_name)
                for col_idx, col_name in enumerate(pivot_std.columns, start=2):
                    value = pivot_std.loc[idx_name, col_name]
                    if pd.notna(value):
                        ws.cell(row=row_offset, column=col_idx, value=value)
                row_offset += 1

        row_offset += 2  # spacing between metrics

    # DETAILED TABLES
    ws.cell(row=row_offset, column=1, value="DETAILED TABLES (By Task)")
    row_offset += 2

    for metric in metrics:
        # Detailed mean
        detailed_mean = detailed_tables.get(metric)
        if detailed_mean is None:
            continue

        ws.cell(row=row_offset, column=1, value=f"Detailed: {metric}")
        row_offset += 1

        # Header
        ws.cell(row=row_offset, column=1, value="Category")
        ws.cell(row=row_offset, column=2, value="Task")
        for col_idx, col_name in enumerate(detailed_mean.columns, start=3):
            ws.cell(row=row_offset, column=col_idx, value=col_name)
        row_offset += 1

        # Write detailed mean rows
        for (cat, task) in detailed_mean.index:
            ws.cell(row=row_offset, column=1, value=cat)
            ws.cell(row=row_offset, column=2, value=task)
            for col_idx, col_name in enumerate(detailed_mean.columns, start=3):
                try:
                    value = detailed_mean.loc[(cat, task), col_name]
                    if pd.notna(value):
                        ws.cell(row=row_offset, column=col_idx, value=value)
                except Exception:
                    pass
            row_offset += 1

        row_offset += 1

        # Detailed std (if present)
        std_key = metric + " (std)"
        detailed_std = detailed_tables.get(std_key)
        if detailed_std is not None and not detailed_std.empty:
            ws.cell(row=row_offset, column=1, value=f"Detailed STD: {metric}")
            row_offset += 1

            ws.cell(row=row_offset, column=1, value="Category")
            ws.cell(row=row_offset, column=2, value="Task")
            for col_idx, col_name in enumerate(detailed_std.columns, start=3):
                ws.cell(row=row_offset, column=col_idx, value=col_name)
            row_offset += 1

            for (cat, task) in detailed_std.index:
                ws.cell(row=row_offset, column=1, value=cat)
                ws.cell(row=row_offset, column=2, value=task)
                for col_idx, col_name in enumerate(detailed_std.columns, start=3):
                    try:
                        value = detailed_std.loc[(cat, task), col_name]
                        if pd.notna(value):
                            ws.cell(row=row_offset, column=col_idx, value=value)
                    except Exception:
                        pass
                row_offset += 1

        row_offset += 2

wb = Workbook()

ws = wb.active
ws.title = "Main"

df["Data thrown out"] = df["n"] - (
    df["mean_count_2_seeds_0_0"]
    + df["mean_count_2_seeds_0_1"]
    + df["mean_count_2_seeds_1_0"]
    + df["mean_count_2_seeds_1_1"]
)
df["Relative Data Thrown Out"] = df["Data thrown out"] / df["n"]

def compute_sanity_check1(row):
    if row["acc_before"] * row["n"] == 0 or (row["1->0"] + row.get("1->1", 0)) == 0:
        return True
    else:
        val1 = row["1->0"] / (row["acc_before"] * row["n"])
        denom = (row["1->0"] + row.get("1->1", 0))
        val2 = row["1->0"] / denom
        return abs(val1 - val2) < 0.0001

def compute_sanity_check2(row):
    diff = abs((row["acc_after"] - row["acc_before"]) - ((row["0->1"] - row["1->0"]) / row["n"]))
    return diff < 0.0001

df["SANITY_CHECK1"] = df.apply(compute_sanity_check1, axis=1)
df["SANITY_CHECK2"] = df.apply(compute_sanity_check2, axis=1)

computed_cols = [
    "Ignore",
    "Number of Correct Answers",
    "num_choices",
    "Abs. Forget.",
    "Rel. Forget.",
    "Abs. Improvement",
    "Rel. Improvement",
    "Adj. Abs. Forgetting (Simplified)",
    "Adj. Abs. Improvement (Simplified)",
    "Data thrown out",
    "Relative Data Thrown Out",
    "SANITY_CHECK1",
    "SANITY_CHECK2",
    "Adj. Abs. Forgetting",
    "Adj. Abs. Improvement",
    "Adj. 2 Seed Abs. Forgetting",
    "Adj. 2 Seed Abs. Improvement",
    "Max Possible Improvement",
    "Max Possible Forgetting",
    "Category",
]

original_cols = list(df.columns)
all_headers = original_cols + [c for c in computed_cols if c not in original_cols]

for col_idx, h in enumerate(all_headers, start=1):
    ws.cell(row=1, column=col_idx, value=h)

col_map = {h: get_column_letter(i + 1) for i, h in enumerate(all_headers)}

for r_idx, row in df.iterrows():
    excel_row = r_idx + 2

    for c_idx, col_name in enumerate(original_cols, start=1):
        val = row[col_name]
        ws.cell(row=excel_row, column=c_idx, value=val)

    A = lambda col: f"{col_map[col]}{excel_row}"

    ws.cell(row=excel_row, column=all_headers.index("Ignore") + 1, value=int(row["Ignore"]))

    task_col = col_map["task"]
    formula_num_correct = (
        f'=IF(OR({task_col}{excel_row}="community|truthfulqa:mc2|0",'
        f' {task_col}{excel_row}="community|truthfulqa_reasoning_letter:mc2|0"), NA(), 1)'
    )
    ws.cell(row=excel_row, column=all_headers.index("Number of Correct Answers") + 1, value=formula_num_correct)

    ws.cell(row=excel_row, column=all_headers.index("num_choices") + 1, value=int(row["num_choices"]))

    ws.cell(row=excel_row, column=all_headers.index("Abs. Forget.") + 1, value=f"={A('1->0')}/{A('n')}")

    ws.cell(row=excel_row, column=all_headers.index("Rel. Forget.") + 1, value=f"={A('1->0')}/({A('n')}*{A('acc_before')})")

    ws.cell(row=excel_row, column=all_headers.index("Abs. Improvement") + 1, value=f"={A('0->1')}/{A('n')}")

    ws.cell(row=excel_row, column=all_headers.index("Rel. Improvement") + 1, value=f"={A('0->1')}/({A('n')}*{A('acc_before')})")

    ws.cell(
        row=excel_row,
        column=all_headers.index("Adj. Abs. Forgetting (Simplified)") + 1,
        value=f"=MAX(({A('1->0')}/{A('n')}) - ((1-{A('acc_before')})*(1-{A('acc_after')})/({A('num_choices')}-1)), 0)",
    )

    ws.cell(
        row=excel_row,
        column=all_headers.index("Adj. Abs. Improvement (Simplified)") + 1,
        value=f"=MAX(({A('0->1')}/{A('n')}) - ((1-{A('acc_before')})*(1-{A('acc_after')})/({A('num_choices')}-1)), 0)",
    )

    ws.cell(
        row=excel_row,
        column=all_headers.index("Data thrown out") + 1,
        value=f"={A('n')}-({A('mean_count_2_seeds_0_0')}+{A('mean_count_2_seeds_0_1')}+{A('mean_count_2_seeds_1_0')}+{A('mean_count_2_seeds_1_1')})",
    )

    dt_col_letter = col_map["Data thrown out"]
    ws.cell(
        row=excel_row,
        column=all_headers.index("Relative Data Thrown Out") + 1,
        value=f"={dt_col_letter}{excel_row}/{A('n')}",
    )

    ws.cell(
        row=excel_row,
        column=all_headers.index("SANITY_CHECK1") + 1,
        value=(
            f'=IF(OR({A("acc_before")}*{A("n")}=0, {A("1->0")}+{A("1->1")}=0), TRUE, '
            f'ABS(({A("1->0")}/({A("acc_before")}*{A("n")})) - ({A("1->0")}/({A("1->0")}+{A("1->1")})))<0.0001)'
        ),
    )

    ws.cell(
        row=excel_row,
        column=all_headers.index("SANITY_CHECK2") + 1,
        value=f"=ABS(({A('acc_after')}-{A('acc_before')}) - (({A('0->1')}-{A('1->0')})/{A('n')}))<0.0001",
    )

    simp_forget_col_letter = col_map["Adj. Abs. Forgetting (Simplified)"]
    ws.cell(
        row=excel_row,
        column=all_headers.index("Adj. Abs. Forgetting") + 1,
        value=f"={simp_forget_col_letter}{excel_row}",
    )

    simp_improve_col_letter = col_map["Adj. Abs. Improvement (Simplified)"]
    ws.cell(
        row=excel_row,
        column=all_headers.index("Adj. Abs. Improvement") + 1,
        value=f"={simp_improve_col_letter}{excel_row}",
    )

    ws.cell(
        row=excel_row,
        column=all_headers.index("Adj. 2 Seed Abs. Forgetting") + 1,
        value=f"={A('mean_count_2_seeds_1_0')}/({A('mean_count_2_seeds_0_0')}+{A('mean_count_2_seeds_0_1')}+{A('mean_count_2_seeds_1_0')}+{A('mean_count_2_seeds_1_1')})",
    )

    ws.cell(
        row=excel_row,
        column=all_headers.index("Adj. 2 Seed Abs. Improvement") + 1,
        value=f"={A('mean_count_2_seeds_0_1')}/({A('mean_count_2_seeds_0_0')}+{A('mean_count_2_seeds_0_1')}+{A('mean_count_2_seeds_1_0')}+{A('mean_count_2_seeds_1_1')})",
    )

    task_cell = A("task")
    xlookup_formula = f'=IFERROR(XLOOKUP(TRIM({task_cell}), Category!$B:$B, Category!$A:$A, "Missing", 0), "Not found")'
    ws.cell(row=excel_row, column=all_headers.index("Category") + 1, value=xlookup_formula)

category_df = pd.DataFrame(category_map, columns=["category", "task"])
ws_cat = wb.create_sheet("Category")
ws_cat.append(["category", "task"])
for _, r in category_df.iterrows():
    ws_cat.append([r["category"], r["task"]])

for group_name, model_pairs in model_groups.items():
    print(f"Processing {group_name}...")
    result = create_pivot_table(df, model_pairs, group_name)
    if result is None:
        continue
    pivot_tables, detailed_tables = result
    write_group_sheet(wb, group_name, pivot_tables, detailed_tables)

wb_main = Workbook()
ws = wb_main.active
ws.title = "Main"

df["Data thrown out"] = df["n"] - (
    df["mean_count_2_seeds_0_0"]
    + df["mean_count_2_seeds_0_1"]
    + df["mean_count_2_seeds_1_0"]
    + df["mean_count_2_seeds_1_1"]
)
df["Relative Data Thrown Out"] = df["Data thrown out"] / df["n"]

df["SANITY_CHECK1"] = df.apply(compute_sanity_check1, axis=1)
df["SANITY_CHECK2"] = df.apply(compute_sanity_check2, axis=1)

computed_cols = [
    "Ignore",
    "Number of Correct Answers",
    "num_choices",
    "Abs. Forget.",
    "Rel. Forget.",
    "Abs. Improvement",
    "Rel. Improvement",
    "Adj. Abs. Forgetting (Simplified)",
    "Adj. Abs. Improvement (Simplified)",
    "Data thrown out",
    "Relative Data Thrown Out",
    "SANITY_CHECK1",
    "SANITY_CHECK2",
    "Adj. Abs. Forgetting",
    "Adj. Abs. Improvement",
    "Adj. 2 Seed Abs. Forgetting",
    "Adj. 2 Seed Abs. Improvement",
    "Adj. Before Accuracy",
    "Adj. After Accuracy",
    "Category",
]

original_cols = list(df.columns)
all_headers = original_cols + [c for c in computed_cols if c not in original_cols]

for col_idx, h in enumerate(all_headers, start=1):
    ws.cell(row=1, column=col_idx, value=h)

col_map = {h: get_column_letter(i + 1) for i, h in enumerate(all_headers)}

for r_idx, row in df.iterrows():
    excel_row = r_idx + 2

    for c_idx, col_name in enumerate(original_cols, start=1):
        val = row[col_name]
        ws.cell(row=excel_row, column=c_idx, value=val)

    A = lambda col: f"{col_map[col]}{excel_row}"

    ws.cell(row=excel_row, column=all_headers.index("Ignore") + 1, value=int(row["Ignore"]))

    task_col = col_map["task"]
    formula_num_correct = (
        f'=IF(OR({task_col}{excel_row}="community|truthfulqa:mc2|0",'
        f' {task_col}{excel_row}="community|truthfulqa_reasoning_letter:mc2|0"), NA(), 1)'
    )
    ws.cell(row=excel_row, column=all_headers.index("Number of Correct Answers") + 1, value=formula_num_correct)

    ws.cell(row=excel_row, column=all_headers.index("num_choices") + 1, value=int(row["num_choices"]))

    ws.cell(row=excel_row, column=all_headers.index("Abs. Forget.") + 1, value=f"={A('1->0')}/{A('n')}")

    ws.cell(row=excel_row, column=all_headers.index("Rel. Forget.") + 1, value=f"={A('1->0')}/({A('n')}*{A('acc_before')})")

    ws.cell(row=excel_row, column=all_headers.index("Abs. Improvement") + 1, value=f"={A('0->1')}/{A('n')}")

    ws.cell(row=excel_row, column=all_headers.index("Rel. Improvement") + 1, value=f"={A('0->1')}/({A('n')}*{A('acc_before')})")

    ws.cell(
        row=excel_row,
        column=all_headers.index("Adj. Before Accuracy") + 1,
        value=f"=(1-{A('acc_before')}) - (1-{A('acc_before')})/({A('n')}-1)"
    )

    ws.cell(
        row=excel_row,
        column=all_headers.index("Adj. After Accuracy") + 1,
        value=f"=(1-{A('acc_after')}) - (1-{A('acc_after')})/({A('n')}-1)"
    )


    ws.cell(
        row=excel_row,
        column=all_headers.index("Adj. Abs. Forgetting (Simplified)") + 1,
        value=f"=MAX(({A('1->0')}/{A('n')}) - ((1-{A('acc_before')})*(1-{A('acc_after')})/({A('num_choices')}-1)), 0)",
    )

    ws.cell(
        row=excel_row,
        column=all_headers.index("Adj. Abs. Improvement (Simplified)") + 1,
        value=f"=MAX(({A('0->1')}/{A('n')}) - ((1-{A('acc_before')})*(1-{A('acc_after')})/({A('num_choices')}-1)), 0)",
    )

    ws.cell(
        row=excel_row,
        column=all_headers.index("Data thrown out") + 1,
        value=f"={A('n')}-({A('mean_count_2_seeds_0_0')}+{A('mean_count_2_seeds_0_1')}+{A('mean_count_2_seeds_1_0')}+{A('mean_count_2_seeds_1_1')})",
    )

    dt_col_letter = col_map["Data thrown out"]
    ws.cell(
        row=excel_row,
        column=all_headers.index("Relative Data Thrown Out") + 1,
        value=f"={dt_col_letter}{excel_row}/{A('n')}",
    )

    ws.cell(
        row=excel_row,
        column=all_headers.index("SANITY_CHECK1") + 1,
        value=(
            f'=IF(OR({A("acc_before")}*{A("n")}=0, {A("1->0")}+{A("1->1")}=0), TRUE, '
            f'ABS(({A("1->0")}/({A("acc_before")}*{A("n")})) - ({A("1->0")}/({A("1->0")}+{A("1->1")})))<0.0001)'
        ),
    )

    ws.cell(
        row=excel_row,
        column=all_headers.index("SANITY_CHECK2") + 1,
        value=f"=ABS(({A('acc_after')}-{A('acc_before')}) - (({A('0->1')}-{A('1->0')})/{A('n')}))<0.0001",
    )

    simp_forget_col_letter = col_map["Adj. Abs. Forgetting (Simplified)"]
    ws.cell(
        row=excel_row,
        column=all_headers.index("Adj. Abs. Forgetting") + 1,
        value=f"={simp_forget_col_letter}{excel_row}",
    )

    simp_improve_col_letter = col_map["Adj. Abs. Improvement (Simplified)"]
    ws.cell(
        row=excel_row,
        column=all_headers.index("Adj. Abs. Improvement") + 1,
        value=f"={simp_improve_col_letter}{excel_row}",
    )

    ws.cell(
        row=excel_row,
        column=all_headers.index("Adj. 2 Seed Abs. Forgetting") + 1,
        value=f"={A('mean_count_2_seeds_1_0')}/({A('mean_count_2_seeds_0_0')}+{A('mean_count_2_seeds_0_1')}+{A('mean_count_2_seeds_1_0')}+{A('mean_count_2_seeds_1_1')})",
    )

    ws.cell(
        row=excel_row,
        column=all_headers.index("Adj. 2 Seed Abs. Improvement") + 1,
        value=f"={A('mean_count_2_seeds_0_1')}/({A('mean_count_2_seeds_0_0')}+{A('mean_count_2_seeds_0_1')}+{A('mean_count_2_seeds_1_0')}+{A('mean_count_2_seeds_1_1')})",
    )

    task_cell = A("task")
    xlookup_formula = f'=IFERROR(XLOOKUP(TRIM({task_cell}), Category!$B:$B, Category!$A:$A, "Missing", 0), "Not found")'
    ws.cell(row=excel_row, column=all_headers.index("Category") + 1, value=xlookup_formula)

category_df = pd.DataFrame(category_map, columns=["category", "task"])
ws_cat = wb_main.create_sheet("Category")
ws_cat.append(["category", "task"])
for _, r in category_df.iterrows():
    ws_cat.append([r["category"], r["task"]])

for sheet_name in wb.sheetnames:
    if sheet_name == "Sheet":  # Skip default sheet
        continue

    source_sheet = wb[sheet_name]
    target_sheet = wb_main.create_sheet(sheet_name)

    for row in source_sheet.iter_rows():
        for cell in row:
            target_sheet[cell.coordinate].value = cell.value

wb_main.save(output_excel_path)
print(f"Done! Created {output_excel_path} with main sheet and pivot tables")

df.to_csv(output_csv_path, index=False)
print(f"Created {output_csv_path}")

os.makedirs(pivot_dir_path, exist_ok=True)

for group_name, model_pairs in model_groups.items():
    result = create_pivot_table(df, model_pairs, group_name)
    if result is None:
        continue

    pivot_tables, detailed_tables = result

    for metric in metrics:
        # mean
        pivot = pivot_tables.get(metric)
        if pivot is not None:
            filename = f"{pivot_dir_path}/{group_name.replace(' ', '_')}_{metric.replace(' ', '_').replace('.', '')}.csv"
            pivot.to_csv(filename, float_format="%.8f", na_rep="")
            print(f"Saved {filename}")

        # var
        var_key = metric + " (var)"
        pivot_var = pivot_tables.get(var_key)
        if pivot_var is not None:
            filename_var = f"{pivot_dir_path}/{group_name.replace(' ', '_')}_{metric.replace(' ', '_').replace('.', '')}_var.csv"
            pivot_var.to_csv(filename_var, float_format="%.10f", na_rep="")
            print(f"Saved {filename_var}")

        # std
        std_key = metric + " (std)"
        pivot_std = pivot_tables.get(std_key)
        if pivot_std is not None:
            filename_std = f"{pivot_dir_path}/{group_name.replace(' ', '_')}_{metric.replace(' ', '_').replace('.', '')}_std.csv"
            pivot_std.to_csv(filename_std, float_format="%.10f", na_rep="")
            print(f"Saved {filename_std}")
